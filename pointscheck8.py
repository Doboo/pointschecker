import openpyxl
from opcua import Client
import configparser
import logging
import os
import time

# 配置日志
logging.basicConfig(filename='opcua_checker.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# 结果数据
results_good = []
results_bad = []
results_not_exist = []
def get_program_path():
  return os.path.dirname(os.path.abspath(sys.argv[0]))
def load_excel(file_path):
    """从Excel文件加载点表"""
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        point_list = []
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                point_list.append(row[0])
        logging.info(f"成功加载Excel文件: {file_path}，共 {len(point_list)} 个点")
        print(f"成功加载Excel文件: {file_path}，共 {len(point_list)} 个点")
        return point_list, workbook, sheet
    except FileNotFoundError:
        logging.error(f"找不到Excel文件: {file_path}")
        print(f"找不到Excel文件: {file_path}")
        return [], None, None
    except Exception as e:
        logging.error(f"加载Excel文件时出错: {e}")
        print(f"加载Excel文件时出错: {e}")
        return [], None, None

def check_points(client, points):
    """检查点是否存在，并读取数据和品质"""
    global results_good, results_bad, results_not_exist
    for point in points:
        try:
            node = client.get_node(point)
            value = node.get_value()
            sss = node.get_data_value()
            quality = sss.StatusCode.name
            result = {"point": point, "exists": True, "value": value, "quality": quality}
            if quality == "Good":
                results_good.append(result)
            else:
                results_bad.append(result)
            logging.debug(f"点 {point} 检查完成，品质: {quality}, 值: {value}")
            print(f"点 {point} 检查完成，品质: {quality}, 值: {value}")
        except Exception as e:
            if "BadNodeIdUnknown" in str(e):
                results_not_exist.append({"point": point, "exists": False, "value": None, "quality": str(e)})
                logging.warning(str(e))
                logging.warning(f"点 {point} 不存在")
                print(f"点 {point} 不存在")
                print(str(e))
            else:
                results_bad.append({"point": point, "exists": False, "value": None, "quality": str(e)})
                logging.error(f"检查点 {point} 时发生其他错误: {e}")
                print(f"检查点 {point} 时发生其他错误: {e}")
                print(str(e))

def write_results_to_excel(output_file):
    """将结果写入Excel文件，并设置第一列宽度"""
    try:
        workbook = openpyxl.Workbook()

        def create_sheet(sheet_name, header, results):
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(header)
            for result in results:
                if sheet_name == "不存在的点":
                    row = [result["point"], "不存在", result["quality"]]
                else:
                    row = [result["point"], "存在", str(result.get("value", "N/A")), result["quality"]]
                sheet.append(row)
            sheet.column_dimensions['A'].width = 25  # 设置第一列宽度为25
            return sheet

        create_sheet("Good品质", ["点名", "存在", "数据值", "品质"], results_good)
        create_sheet("Bad品质", ["点名", "存在", "数据值", "品质"], results_bad)
        create_sheet("不存在的点", ["点名", "存在", "原因"], results_not_exist)

        workbook.save(output_file)
        logging.info(f"结果已保存到: {output_file}")
        print(f"结果已保存到: {output_file}")
    except Exception as e:
        logging.error(f"写入Excel文件 {output_file} 时出错: {e}")
        print(f"写入Excel文件 {output_file} 时出错: {e}")

def main():
    """主函数"""
    config = configparser.ConfigParser()
    try:
        config.read('config.ini', encoding='utf-8')
    except Exception as e:
        logging.error(f"读取配置文件 config.ini 出错: {e}")
        print(f"读取配置文件 config.ini 出错: {e}")
        input("按回车键退出...")
        return

    opcua_server_url = config.get('OPCUA', 'server_url')
    opcua_server_name = config.get('OPCUA', 'server_name', fallback="Unknown Server")
    print(f"OPC UA Server Info:")
    print(f"  Name: {opcua_server_name}")
    print(f"  URL: {opcua_server_url}")
    logging.info(f"尝试连接到OPC UA服务器: {opcua_server_name}({opcua_server_url})")

    current_dir = get_program_path()
    
    excel_file_path = os.path.join(current_dir, "points.xlsx")

    points, _, _ = load_excel(excel_file_path)
    if not points:
        input("加载点表失败，按回车键退出...")
        return

    client = Client(opcua_server_url)
    try:
        print("正在连接到OPC UA服务器...")
        client.connect()
        print("成功连接到OPC UA服务器")
        logging.info(f"成功连接到OPC UA服务器: {opcua_server_url}")

        check_points(client, points)

        output_file_path = os.path.join(current_dir, "result.xlsx")

        write_results_to_excel(output_file_path)

    except Exception as e:
        logging.error(f"连接或操作OPC UA服务器时出错: {e}")
        print(f"连接OPC UA服务器出错: {e}")
        print("请检查服务器是否可用，网络连接是否正常。")
        input("按回车键退出...")
        return
    finally:
        try:
            client.disconnect()
            logging.info("已断开与OPC UA服务器的连接")
            print("已断开与OPC UA服务器的连接")
        except Exception as e:
            logging.error(f"断开连接时出错: {e}")
            print(f"断开连接时出错: {e}")

    input("按回车键退出...")

if __name__ == "__main__":
    main()